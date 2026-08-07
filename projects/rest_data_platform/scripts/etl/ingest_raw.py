#!/usr/bin/env python3
"""
Layer-1: Quelltreuer Import nach imc_source_snapshots + imc_source_raw_rows.

Beispiel:
  python ingest_raw.py --source-id a1000001-0001-4001-8001-000000000001 \\
    --file "C:/.../4COffshore - Offshore Wind Farm Database ..._SAMPLE.xlsx"
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import psycopg
from dotenv import load_dotenv

from config import FOURC_SOURCE_PROFILES, SOURCE_4C_WINDFARM, get_database_url

load_dotenv()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def json_safe(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def infer_profile_name(path: Path) -> str | None:
    name = path.name.lower()
    if "pop" in name:
        return "pop"
    if "wind turbine" in name or "turbine database" in name:
        return "turbine"
    if "vessel" in name or "ports intelligence" in name or "vpi" in name:
        return "vpi"
    if "transmission" in name or "cables database" in name:
        return "transmission"
    if "interconnector" in name:
        return "interconnectors"
    if "wind farm" in name or "windfarm" in name:
        return "windfarm"
    return None


def get_profile(profile_name: str | None) -> dict[str, Any]:
    if profile_name is None:
        return {
            "source_id": SOURCE_4C_WINDFARM,
            "default_header_row": 1,
            "key_fields": [
                "WindfarmId",
                "VesselId",
                "PlatformID",
                "WindfarmEventId",
            ],
        }
    try:
        return FOURC_SOURCE_PROFILES[profile_name]
    except KeyError as exc:
        known = ", ".join(sorted(FOURC_SOURCE_PROFILES))
        raise ValueError(f"Unbekanntes Profil '{profile_name}'. Bekannt: {known}") from exc


def normalize_header(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def ext_key_from_payload(payload: dict, key_fields: list[str]) -> str | None:
    for key in key_fields:
        value = payload.get(key)
        if value not in (None, ""):
            return str(value)
    return None


def load_xlsx_rows(
    path: Path,
    profile: dict[str, Any],
) -> list[tuple[str, int, str | None, dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out: list[tuple[str, int, str | None, dict]] = []
    default_header_row = int(profile.get("default_header_row", 1))
    sheet_header_rows = profile.get("sheet_header_rows", {})
    include_sheets = set(profile.get("include_sheets", []) or [])
    key_fields = list(profile.get("key_fields", []))

    for sheet_name in wb.sheetnames:
        if include_sheets and sheet_name not in include_sheets:
            continue

        ws = wb[sheet_name]
        header_row_num = int(
            sheet_header_rows.get(sheet_name, default_header_row)
            if isinstance(sheet_header_rows, dict)
            else default_header_row
        )
        header_row = next(
            ws.iter_rows(min_row=header_row_num, max_row=header_row_num),
            None,
        )
        if not header_row:
            continue
        headers = [normalize_header(c.value) for c in header_row]
        for idx, row in enumerate(
            ws.iter_rows(min_row=header_row_num + 1, values_only=True),
            start=header_row_num + 1,
        ):
            if not any(row):
                continue
            payload = {str(h): json_safe(v) for h, v in zip(headers, row) if h}
            if not payload:
                continue
            ext_key = ext_key_from_payload(payload, key_fields)
            rows_out.append((sheet_name, idx, ext_key, payload))
    wb.close()
    return rows_out


def upsert_snapshot(
    conn: psycopg.Connection,
    source_id: str,
    source_file: str,
    file_hash: str,
    row_count: int,
) -> str:
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE public.imc_data_sources
            SET snapshot_date = %s, file_hash = %s
            WHERE source_id = %s
            """,
            (date.today(), file_hash, source_id),
        )
        cur.execute(
            """
            INSERT INTO public.imc_source_snapshots
                (source_id, source_file, file_hash, row_count, notes)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (source_id, file_hash) DO UPDATE SET
                row_count = EXCLUDED.row_count,
                ingested_at = now()
            RETURNING snapshot_id
            """,
            (source_id, source_file, file_hash, row_count, "ingest_raw.py"),
        )
        snapshot_id = cur.fetchone()[0]
    conn.commit()
    return str(snapshot_id)


def insert_raw_rows(
    conn: psycopg.Connection,
    snapshot_id: str,
    source_id: str,
    rows: list[tuple[str, int, str | None, dict]],
) -> int:
    with conn.cursor() as cur:
        cur.execute(
            "DELETE FROM public.imc_source_raw_rows WHERE snapshot_id = %s",
            (snapshot_id,),
        )
        cur.executemany(
            """
            INSERT INTO public.imc_source_raw_rows
                (snapshot_id, source_id, sheet_name, source_row_num, ext_record_key, payload)
            VALUES (%s, %s, %s, %s, %s, %s::jsonb)
            """,
            [
                (snapshot_id, source_id, sheet, row_num, ext_key, json.dumps(payload))
                for sheet, row_num, ext_key, payload in rows
            ],
        )
    conn.commit()
    return len(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="IMC raw mirror ingest")
    parser.add_argument("--source-id")
    parser.add_argument(
        "--profile",
        choices=sorted(FOURC_SOURCE_PROFILES),
        help="4C source profile. If omitted, inferred from filename when possible.",
    )
    parser.add_argument("--file", required=True, type=Path)
    args = parser.parse_args()

    path = args.file.resolve()
    if not path.exists():
        print(f"Datei nicht gefunden: {path}", file=sys.stderr)
        return 1

    profile_name = args.profile or infer_profile_name(path)
    profile = get_profile(profile_name)
    source_id = args.source_id or str(profile["source_id"])

    file_hash = sha256_file(path)
    rows = load_xlsx_rows(path, profile)
    print(
        f"Gelesen: {len(rows)} Zeilen aus {path.name} "
        f"(profile={profile_name or 'default'}, source_id={source_id}, "
        f"sha256={file_hash[:12]}...)"
    )

    with psycopg.connect(get_database_url()) as conn:
        snapshot_id = upsert_snapshot(
            conn, source_id, str(path), file_hash, len(rows)
        )
        n = insert_raw_rows(conn, snapshot_id, source_id, rows)

    print(f"Import OK: snapshot_id={snapshot_id}, rows={n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
